from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_PATH = Path(__file__).with_name("task_progress_appendix.xlsx")

HEADERS = [
    "task_id",
    "assignment_document_number",
    "task_description",
    "assigned_unit",
    "coordinating_units",
    "assigned_date",
    "deadline",
    "status",
    "completion_percentage",
    "result_summary",
    "evidence_document",
    "delay_reason",
    "next_action",
]

DESCRIPTIONS = {
    "task_id": "Mã định danh duy nhất của nhiệm vụ.",
    "assignment_document_number": "Số, ký hiệu văn bản giao nhiệm vụ.",
    "task_description": "Nội dung nhiệm vụ được giao.",
    "assigned_unit": "Đơn vị chủ trì thực hiện.",
    "coordinating_units": "Các đơn vị phối hợp; phân tách bằng dấu chấm phẩy (;).",
    "assigned_date": "Ngày giao nhiệm vụ, định dạng ngày/tháng/năm.",
    "deadline": "Thời hạn hoàn thành, định dạng ngày/tháng/năm.",
    "status": "Trạng thái chuẩn, chọn từ danh sách có sẵn.",
    "completion_percentage": "Tỷ lệ hoàn thành từ 0% đến 100%.",
    "result_summary": "Tóm tắt kết quả thực hiện.",
    "evidence_document": "Số văn bản, đường dẫn hoặc tài liệu chứng minh kết quả.",
    "delay_reason": "Lý do chậm tiến độ; bắt buộc về nghiệp vụ khi trạng thái là delayed.",
    "next_action": "Công việc hoặc hành động tiếp theo.",
}

STATUS_VALUES = [
    "not_started",
    "in_progress",
    "completed",
    "delayed",
    "on_hold",
    "cancelled",
]


def build_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "task_progress"
    sheet.append(HEADERS)
    sheet.append([None] * len(HEADERS))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(DESCRIPTIONS[cell.value], "Codex")

    sheet.row_dimensions[1].height = 36
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:M1001"

    widths = {
        "A": 16, "B": 28, "C": 48, "D": 24, "E": 32,
        "F": 16, "G": 16, "H": 18, "I": 23, "J": 40,
        "K": 36, "L": 36, "M": 36,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in range(2, 1002):
        sheet[f"F{row}"].number_format = "dd/mm/yyyy"
        sheet[f"G{row}"].number_format = "dd/mm/yyyy"
        sheet[f"I{row}"].number_format = "0%"
        for column in ("C", "E", "J", "K", "L", "M"):
            sheet[f"{column}{row}"].alignment = Alignment(vertical="top", wrap_text=True)

    lists = workbook.create_sheet("_lists")
    for index, status in enumerate(STATUS_VALUES, start=1):
        lists.cell(index, 1, status)
    lists.sheet_state = "hidden"

    status_validation = DataValidation(
        type="list", formula1="'_lists'!$A$1:$A$6", allow_blank=True
    )
    status_validation.error = "Hãy chọn một trạng thái trong danh sách."
    status_validation.errorTitle = "Trạng thái không hợp lệ"
    status_validation.prompt = "Chọn trạng thái thực hiện nhiệm vụ."
    status_validation.promptTitle = "Trạng thái"
    status_validation.showErrorMessage = True
    status_validation.showInputMessage = True
    sheet.add_data_validation(status_validation)
    status_validation.add("H2:H1001")

    percentage_validation = DataValidation(
        type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True
    )
    percentage_validation.error = "Nhập tỷ lệ trong khoảng 0% đến 100%."
    percentage_validation.errorTitle = "Tỷ lệ không hợp lệ"
    percentage_validation.showErrorMessage = True
    sheet.add_data_validation(percentage_validation)
    percentage_validation.add("I2:I1001")

    status_colors = {
        "not_started": "D9EAF7",
        "in_progress": "FFF2CC",
        "completed": "C6E0B4",
        "delayed": "F4CCCC",
        "on_hold": "D9D2E9",
        "cancelled": "D9D9D9",
    }
    for status, color in status_colors.items():
        sheet.conditional_formatting.add(
            "H2:H1001",
            FormulaRule(formula=[f'$H2="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )

    table = Table(displayName="TaskProgressTable", ref="A1:M2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.sheet_view.showGridLines = False
    return workbook


if __name__ == "__main__":
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(OUTPUT_PATH)
    print(OUTPUT_PATH)
