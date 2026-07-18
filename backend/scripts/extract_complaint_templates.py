"""Extract complaint-report templates from the official TT06 DOCX appendix."""

from __future__ import annotations

import argparse
import shutil
from dataclasses import dataclass
from pathlib import Path

from docx import Document
from docx.oxml.ns import qn
from docx.table import Table, _Cell
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class FormSpec:
    code: str
    next_code: str
    output_name: str


FORM_SPECS = (
    FormSpec("01/TCD", "02/XLD", "bieu_01_TCD_tiep_cong_dan.xlsx"),
    FormSpec("02/XLD", "03/GQKN", "bieu_02_XLD_tiep_nhan_xu_ly_don.xlsx"),
    FormSpec("03/GQKN", "04/GQTC", "bieu_03_GQKN_giai_quyet_khieu_nai.xlsx"),
    FormSpec("04/GQTC", "05/KQTH", "bieu_04_GQTC_giai_quyet_to_cao.xlsx"),
    FormSpec("05/KQTH", "06/QLKNTC", "bieu_05_KQTH_ket_qua_thi_hanh.xlsx"),
    FormSpec("06/QLKNTC", "Biểu số: 01/PCTN", "bieu_06_QLKNTC_quan_ly_nha_nuoc.xlsx"),
)


def _element_text(element: object, document: Document) -> str:
    if element.tag == qn("w:p"):
        return Paragraph(element, document._body).text.strip()  # noqa: SLF001
    return ""


def _find_paragraph_block(document: Document, marker: str, start: int = 0) -> int:
    elements = list(document.element.body.iterchildren())
    for index in range(start, len(elements)):
        text = _element_text(elements[index], document)
        if text.startswith(marker):
            return index
    raise ValueError(f"marker not found in source DOCX: {marker}")


def _find_table_block(document: Document, start: int, end: int) -> int:
    elements = list(document.element.body.iterchildren())
    for index in range(start, end):
        if elements[index].tag == qn("w:tbl"):
            return index
    raise ValueError(f"table not found between body blocks {start} and {end}")


def _paragraphs_between(document: Document, start: int, end: int) -> list[str]:
    elements = list(document.element.body.iterchildren())
    return [text for element in elements[start:end] if (text := _element_text(element, document))]


def _table_from_element(document: Document, element: object) -> Table:
    return Table(element, document._body)  # noqa: SLF001


def _grid_span(tc: object) -> int:
    grid_span = tc.tcPr.gridSpan
    return int(grid_span.val) if grid_span is not None else 1


def _vertical_merge(tc: object) -> str | None:
    merge = tc.tcPr.vMerge
    if merge is None:
        return None
    return "continue" if merge.val is None else str(merge.val)


def _grid_before(tr: object) -> int:
    value = tr.xpath("./w:trPr/w:gridBefore/@w:val")
    return int(value[0]) if value else 0


def _merge_if_needed(
    sheet: object, start_row: int, end_row: int, start_col: int, span: int
) -> None:
    end_col = start_col + span - 1
    if end_row > start_row or end_col > start_col:
        sheet.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col,
        )


def _write_word_table(sheet: object, table: Table, start_row: int) -> tuple[int, int]:
    grid = table._tbl.tblGrid.gridCol_lst  # noqa: SLF001
    column_count = len(grid)
    active_vertical: dict[tuple[int, int], int] = {}
    thin = Side(style="thin", color="000000")

    for relative_row, tr in enumerate(table._tbl.tr_lst):  # noqa: SLF001
        excel_row = start_row + relative_row
        column = _grid_before(tr) + 1
        continued: set[tuple[int, int]] = set()

        for tc in tr.tc_lst:
            span = _grid_span(tc)
            key = (column, span)
            merge_type = _vertical_merge(tc)
            text = _Cell(tc, table).text.strip()

            if merge_type == "continue":
                continued.add(key)
                if key not in active_vertical:
                    raise ValueError(
                        f"orphan vertical merge at row {relative_row + 1}, column {column}"
                    )
            else:
                if key in active_vertical:
                    merge_start = active_vertical.pop(key)
                    _merge_if_needed(sheet, merge_start, excel_row - 1, column, span)

                cell = sheet.cell(excel_row, column, text)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                if merge_type == "restart":
                    active_vertical[key] = excel_row
                    continued.add(key)
                else:
                    _merge_if_needed(sheet, excel_row, excel_row, column, span)

            column += span

        for key in list(active_vertical):
            if key not in continued and active_vertical[key] < excel_row:
                merge_start = active_vertical.pop(key)
                _merge_if_needed(sheet, merge_start, excel_row - 1, key[0], key[1])

    last_row = start_row + len(table.rows) - 1
    for (column, span), merge_start in active_vertical.items():
        _merge_if_needed(sheet, merge_start, last_row, column, span)

    for relative_row in range(len(table.rows)):
        sheet.row_dimensions[start_row + relative_row].height = 34

    for column, grid_column in enumerate(grid, start=1):
        width_twips = int(grid_column.w or 1440)
        sheet.column_dimensions[get_column_letter(column)].width = max(
            4.0, min(24.0, width_twips / 180.0)
        )

    return last_row, column_count


def _write_form_workbook(
    document: Document,
    spec: FormSpec,
    destination: Path,
) -> None:
    elements = list(document.element.body.iterchildren())
    start = _find_paragraph_block(document, spec.code)
    end = _find_paragraph_block(document, spec.next_code, start + 1)
    table_index = _find_table_block(document, start, end)
    headings = _paragraphs_between(document, start, table_index)
    notes = _paragraphs_between(document, table_index + 1, end)
    table = _table_from_element(document, elements[table_index])

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.code.replace("/", "_")
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    column_count = len(table._tbl.tblGrid.gridCol_lst)  # noqa: SLF001
    for row, text in enumerate(headings, start=1):
        sheet.cell(row, 1, text)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=column_count)
        sheet.cell(row, 1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet.cell(row, 1).font = Font(bold=row <= 2, size=12 if row <= 2 else 10)

    table_start = len(headings) + 2
    table_end, column_count = _write_word_table(sheet, table, table_start)
    for cell in sheet[table_start]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True, size=9)

    note_row = table_end + 2
    for text in notes:
        sheet.cell(note_row, 1, text)
        sheet.merge_cells(
            start_row=note_row,
            start_column=1,
            end_row=note_row,
            end_column=column_count,
        )
        sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(note_row, 1).font = Font(size=9, italic=text.startswith("Hướng dẫn"))
        sheet.row_dimensions[note_row].height = max(18, min(90, 15 * (1 + len(text) // 140)))
        note_row += 1

    sheet.freeze_panes = sheet.cell(table_start, 2)
    sheet.print_area = f"A1:{get_column_letter(column_count)}{note_row - 1}"
    sheet.oddFooter.center.text = "Nguồn: Phụ lục Thông tư 06/2025/TT-TTCP"
    workbook.properties.title = f"{spec.code} - Thông tư 06/2025/TT-TTCP"
    workbook.properties.subject = "Biểu mẫu được tách nguyên cấu trúc từ DOCX nguồn"
    workbook.properties.creator = "VAIC2026 dataset extraction"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _extract_mau_02(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    document = Document(destination)
    body = document.element.body
    elements = list(body.iterchildren())
    start = _find_paragraph_block(document, "Mẫu số 02.")
    end = _find_paragraph_block(document, "Mẫu số 03.", start + 1)

    for index, element in enumerate(elements):
        if element.tag == qn("w:sectPr"):
            continue
        if index < start or index >= end:
            body.remove(element)

    document.core_properties.title = "Mẫu số 02 - Thông tư 06/2025/TT-TTCP"
    document.core_properties.subject = (
        "Đề cương báo cáo tiếp công dân, giải quyết khiếu nại, tố cáo"
    )
    document.save(destination)


def extract(source: Path, output_directory: Path) -> None:
    if not source.is_file():
        raise FileNotFoundError(source)
    document = Document(source)
    _extract_mau_02(source, output_directory / "mau_02_bao_cao_kntc.docx")
    for spec in FORM_SPECS:
        _write_form_workbook(document, spec, output_directory / spec.output_name)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output_directory", type=Path)
    args = parser.parse_args()
    extract(args.source.resolve(), args.output_directory.resolve())


if __name__ == "__main__":
    main()
