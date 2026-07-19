"""Deterministic XLSX extraction for the three-source population bundle."""

from __future__ import annotations

import calendar
import hashlib
import re
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook

from app.operational_data.models import (
    DataClassification,
    IssueSeverity,
    OrganizationMetadata,
    PopulationExtractedSource,
    PopulationSourceRole,
    ReportingPeriod,
    ReportingPeriodKind,
    SourceProvenance,
    ValidationIssue,
)

ROLE_FIELDS: dict[PopulationSourceRole, frozenset[str]] = {
    PopulationSourceRole.OPENING_BALANCE: frozenset(
        {"population_opening", "temporary_opening"}
    ),
    PopulationSourceRole.CIVIL_STATUS: frozenset(
        {
            "birth_registered",
            "birth_local_resident",
            "death_registered",
            "death_local_resident",
        }
    ),
    PopulationSourceRole.RESIDENCE_MOVEMENT: frozenset(
        {"permanent_in", "permanent_out", "temporary_new", "temporary_removed"}
    ),
}


class PopulationExtractionError(ValueError):
    """Raised with structured issues when a population workbook is unsafe to use."""

    def __init__(self, message: str, issues: list[ValidationIssue] | None = None) -> None:
        super().__init__(message)
        self.issues = issues or []


def _key(value: object) -> str:
    text = unicodedata.normalize("NFC", str(value or "")).strip().casefold()
    return re.sub(r"\s+", " ", text).rstrip(":")


def _integer(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip().replace(" ", "")
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    return None


def _issue(code: str, message: str, field: str | None = None) -> ValidationIssue:
    return ValidationIssue(
        code=code,
        message=message,
        severity=IssueSeverity.ERROR,
        field_path=field,
    )


class PopulationWorkbookExtractor:
    """Recognize source roles by workbook structure and reconcile summary/detail counts."""

    def extract(
        self,
        source_path: Path,
        expected_role: PopulationSourceRole | None = None,
        source_filename: str | None = None,
    ) -> PopulationExtractedSource:
        if source_path.suffix.lower() != ".xlsx":
            raise PopulationExtractionError("population bundle sources must be XLSX files")
        filename = source_filename or source_path.name
        cached = self._read_workbook(source_path, data_only=True)
        formulas = self._read_workbook(source_path, data_only=False)
        role = self._detect_role(formulas)
        if expected_role is not None and role is not expected_role:
            raise PopulationExtractionError(
                f"workbook role is {role.value}, expected {expected_role.value}",
                [
                    _issue(
                        "POP_SOURCE_ROLE_MISMATCH",
                        "Vai trò nguồn phát hiện được không khớp trường upload.",
                        "source_role",
                    )
                ],
            )

        period = self._period(cached)
        organization = self._organization(cached)
        all_text = "\n".join(
            str(value)
            for rows in cached.values()
            for row in rows
            for value in row
            if value is not None
        )
        classification = (
            DataClassification.SYNTHETIC_TEST_DATA
            if "SYNTHETIC_TEST_DATA" in all_text.upper()
            else DataClassification.OFFICIAL_CANDIDATE
        )
        checksum = hashlib.sha256(source_path.read_bytes()).hexdigest()
        summaries, summary_locations = self._summary_values(cached)
        detail_counts, detail_records, detail_sheet = self._detail_counts(role, cached)
        values: dict[str, int] = {}
        provenance: dict[str, list[SourceProvenance]] = {}
        warnings: list[ValidationIssue] = []

        for field in ROLE_FIELDS[role]:
            summary_value = summaries.get(field)
            detail_value = detail_counts.get(field)
            if summary_value is not None and detail_value is not None:
                if summary_value != detail_value:
                    raise PopulationExtractionError(
                        f"summary/detail mismatch for {field}",
                        [
                            _issue(
                                "POP_SUMMARY_DETAIL_MISMATCH",
                                (
                                    f"Giá trị tổng hợp {field}={summary_value} không khớp "
                                    f"số đếm chi tiết {detail_value}."
                                ),
                                field,
                            )
                        ],
                    )
                value = summary_value
            elif detail_value is not None:
                value = detail_value
                warnings.append(
                    ValidationIssue(
                        code="POP_SUMMARY_VALUE_DERIVED_FROM_DETAIL",
                        message=f"{field} được tính lại từ bảng chi tiết vì thiếu cached value.",
                        severity=IssueSeverity.WARNING,
                        field_path=field,
                    )
                )
            elif summary_value is not None:
                value = summary_value
            else:
                raise PopulationExtractionError(
                    f"required population indicator is missing: {field}",
                    [_issue("POP_REQUIRED_FIELD_MISSING", "Thiếu chỉ tiêu bắt buộc.", field)],
                )
            if value < 0:
                raise PopulationExtractionError(
                    f"population indicator must be non-negative: {field}",
                    [_issue("POP_NEGATIVE_VALUE", "Chỉ tiêu phải là số nguyên không âm.", field)],
                )
            values[field] = value
            location = summary_locations.get(field, (detail_sheet, None))
            provenance[field] = [
                SourceProvenance(
                    source_file_id=filename,
                    source_sha256=checksum,
                    source_role=role,
                    sheet_name=location[0],
                    row_number=location[1],
                    extraction_method="openpyxl_summary_detail_reconciliation",
                    evidence=field,
                )
            ]

        if classification is DataClassification.OFFICIAL_CANDIDATE:
            warnings.append(
                ValidationIssue(
                    code="UNAPPROVED_OFFICIAL_MAPPING",
                    message="Nguồn không có nhãn synthetic; cần cán bộ xác nhận trước khi sử dụng.",
                    severity=IssueSeverity.WARNING,
                )
            )
        return PopulationExtractedSource(
            role=role,
            source_filename=filename,
            source_sha256=checksum,
            reporting_period=period,
            organization=organization,
            classification=classification,
            values=values,
            provenance=provenance,
            sheet_names=list(cached),
            detail_counts=detail_counts,
            detail_record_count=detail_records,
            extraction_warnings=warnings,
        )

    @staticmethod
    def _read_workbook(path: Path, *, data_only: bool) -> dict[str, list[list[Any]]]:
        workbook = load_workbook(path, data_only=data_only, read_only=True)
        try:
            return {
                sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
                for sheet in workbook.worksheets
            }
        finally:
            workbook.close()

    @staticmethod
    def _detect_role(workbook: dict[str, list[list[Any]]]) -> PopulationSourceRole:
        codes = {
            str(row[0]).strip()
            for rows in workbook.values()
            for row in rows
            if row and isinstance(row[0], str)
        }
        matches = [role for role, fields in ROLE_FIELDS.items() if fields.issubset(codes)]
        if len(matches) != 1:
            raise PopulationExtractionError(
                "population source role could not be determined uniquely",
                [
                    _issue(
                        "POP_SOURCE_ROLE_UNDETERMINED",
                        "Không thể xác định duy nhất vai trò workbook từ mã chỉ tiêu và cấu trúc.",
                        "source_role",
                    )
                ],
            )
        return matches[0]

    @staticmethod
    def _period(workbook: dict[str, list[list[Any]]]) -> ReportingPeriod:
        candidates: list[str] = []
        for rows in workbook.values():
            for row in rows:
                for index, value in enumerate(row):
                    if _key(value) == "kỳ báo cáo" and index + 1 < len(row):
                        candidates.insert(0, str(row[index + 1]))
                    elif value is not None:
                        candidates.append(str(value))
        for text in candidates:
            match = re.search(r"tháng\s*(\d{1,2})\s*[/_-]\s*(\d{4})", text, re.I)
            if match:
                month, year = int(match.group(1)), int(match.group(2))
                return ReportingPeriod(
                    kind=ReportingPeriodKind.MONTH,
                    start=date(year, month, 1),
                    end=date(year, month, calendar.monthrange(year, month)[1]),
                    label=f"Tháng {month:02d}/{year}",
                )
        raise PopulationExtractionError(
            "reporting period could not be determined from workbook content",
            [_issue("POP_REPORTING_PERIOD_MISSING", "Không xác định được kỳ báo cáo.")],
        )

    @staticmethod
    def _organization(workbook: dict[str, list[list[Any]]]) -> OrganizationMetadata:
        for rows in workbook.values():
            for row in rows:
                for index, value in enumerate(row):
                    if _key(value) == "đơn vị tổng hợp" and index + 1 < len(row):
                        name = str(row[index + 1] or "").strip()
                        if name:
                            return OrganizationMetadata(
                                organization_id=str(uuid5(NAMESPACE_URL, _key(name))),
                                organization_name=name,
                            )
        raise PopulationExtractionError(
            "reporting organization could not be determined",
            [_issue("POP_REPORTING_ORGANIZATION_MISSING", "Thiếu đơn vị tổng hợp.")],
        )

    @staticmethod
    def _summary_values(
        workbook: dict[str, list[list[Any]]],
    ) -> tuple[dict[str, int], dict[str, tuple[str, int]]]:
        values: dict[str, int] = {}
        locations: dict[str, tuple[str, int]] = {}
        canonical = set().union(*ROLE_FIELDS.values())
        for sheet_name, rows in workbook.items():
            for row_number, row in enumerate(rows, start=1):
                if not row or str(row[0] or "").strip() not in canonical:
                    continue
                value = _integer(row[2] if len(row) > 2 else None)
                if value is not None:
                    values[str(row[0]).strip()] = value
                    locations[str(row[0]).strip()] = (sheet_name, row_number)
        return values, locations

    def _detail_counts(
        self,
        role: PopulationSourceRole,
        workbook: dict[str, list[list[Any]]],
    ) -> tuple[dict[str, int], int, str | None]:
        if role is PopulationSourceRole.OPENING_BALANCE:
            return self._opening_counts(workbook)
        if role is PopulationSourceRole.CIVIL_STATUS:
            return self._civil_counts(workbook)
        return self._movement_counts(workbook)

    @staticmethod
    def _find_header(
        workbook: dict[str, list[list[Any]]], required: set[str]
    ) -> tuple[str, list[list[Any]], int, dict[str, int]]:
        for sheet_name, rows in workbook.items():
            for index, row in enumerate(rows):
                headers = {
                    _key(value): column
                    for column, value in enumerate(row)
                    if value is not None
                }
                if required.issubset(headers):
                    return sheet_name, rows, index, headers
        raise PopulationExtractionError(
            "required detail table was not found",
            [_issue("POP_DETAIL_TABLE_MISSING", "Không tìm thấy bảng chi tiết bắt buộc.")],
        )

    def _opening_counts(
        self, workbook: dict[str, list[list[Any]]]
    ) -> tuple[dict[str, int], int, str]:
        sheet, rows, header_index, headers = self._find_header(
            workbook, {"mã địa bàn", "tổng thường trú", "tổng tạm trú"}
        )
        permanent = temporary = records = 0
        for row in rows[header_index + 1 :]:
            identity = str(row[headers["mã địa bàn"]] or "").strip()
            if _key(identity) == "tổng cộng":
                break
            if not identity:
                continue
            permanent_value = _integer(row[headers["tổng thường trú"]])
            temporary_value = _integer(row[headers["tổng tạm trú"]])
            if permanent_value is None or temporary_value is None:
                raise PopulationExtractionError(
                    "opening-balance detail contains a non-integer value",
                    [_issue("POP_DETAIL_VALUE_INVALID", "Bảng địa bàn có giá trị không hợp lệ.")],
                )
            permanent += permanent_value
            temporary += temporary_value
            records += 1
        return {
            "population_opening": permanent,
            "temporary_opening": temporary,
        }, records, sheet

    def _civil_counts(
        self, workbook: dict[str, list[list[Any]]]
    ) -> tuple[dict[str, int], int, str]:
        sheet, rows, header_index, headers = self._find_header(
            workbook, {"mã sự kiện", "loại sự kiện", "thuộc dân cư thường trú xã"}
        )
        event_counts: Counter[str] = Counter()
        local_counts: Counter[str] = Counter()
        records = 0
        for row in rows[header_index + 1 :]:
            if not str(row[headers["mã sự kiện"]] or "").strip():
                continue
            event = str(row[headers["loại sự kiện"]] or "").strip().upper()
            if event not in {"KHAI_SINH", "KHAI_TU"}:
                continue
            event_counts[event] += 1
            if _key(row[headers["thuộc dân cư thường trú xã"]]) == "có":
                local_counts[event] += 1
            records += 1
        return {
            "birth_registered": event_counts["KHAI_SINH"],
            "birth_local_resident": local_counts["KHAI_SINH"],
            "death_registered": event_counts["KHAI_TU"],
            "death_local_resident": local_counts["KHAI_TU"],
        }, records, sheet

    def _movement_counts(
        self, workbook: dict[str, list[list[Any]]]
    ) -> tuple[dict[str, int], int, str]:
        sheet, rows, header_index, headers = self._find_header(
            workbook, {"mã biến động", "loại biến động"}
        )
        counts: Counter[str] = Counter()
        accepted = {"THUONG_TRU_DEN", "THUONG_TRU_DI", "TAM_TRU_MOI", "TAM_TRU_KET_THUC"}
        records = 0
        for row in rows[header_index + 1 :]:
            if not str(row[headers["mã biến động"]] or "").strip():
                continue
            movement = str(row[headers["loại biến động"]] or "").strip().upper()
            if movement not in accepted:
                continue
            counts[movement] += 1
            records += 1
        return {
            "permanent_in": counts["THUONG_TRU_DEN"],
            "permanent_out": counts["THUONG_TRU_DI"],
            "temporary_new": counts["TAM_TRU_MOI"],
            "temporary_removed": counts["TAM_TRU_KET_THUC"],
        }, records, sheet
